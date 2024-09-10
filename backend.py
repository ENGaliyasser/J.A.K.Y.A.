from front import Ui_MainWindow
from PyQt5 import QtCore, QtGui, QtWidgets, QtTest
import sys
from PyQt5.QtWidgets import QMessageBox, QFileDialog, QApplication, QMainWindow, QLineEdit, QVBoxLayout, QWidget, QTabWidget
import openpyxl
from PyQt5.QtCore import QTimer, pyqtSignal, QObject, QThread
import os
import sys
from PyQt5 import QtWidgets
from updater import Updater
import time
from datetime import date
import access_sheet
import gspread
import createfiles
import SCANNER
import pieChart

# Global Variables
CURRENT_MODE = None       # Variable that holds the current mode
password = "jakya2024"    # Variable to store the password
SCANNER_CONNECTED = False # Variable to check if the scanner is connected

# Functions
def check_password(entered_pass):
    """Checks if the entered password is correct.

    Args:
        entered_pass: string of the password to check.

    Returns:
        True if the password is correct, False otherwise.
    """
    if entered_pass == password:
        return True
    else:
        return False


class FocusRedirector(QObject):
    """
    FocusRedirector is a class that ensures a QLineEdit always 
    receives focus whenever certain events occur within the 
    parent widget it monitors.

    Attributes
    ----------
    target : QWidget
        The target widget that should always receive focus.

    Methods
    -------
    eventFilter(obj, event)
        Monitors and handles events, redirecting focus to the target widget 
        when specific events occur.
    """
    def __init__(self, target):
        super().__init__()
        self.target = target

    def eventFilter(self, obj, event):
        """
        Monitors and handles events, redirecting focus to the target widget 
        when specific events occur.

        Parameters
        ----------
        obj : QObject
            The object for which the event is being filtered.
        event : QEvent
            The event to be filtered and possibly handled.

        Returns
        -------
        bool
            True if the event should be filtered out, False otherwise.
        """
        if event.type() in (event.FocusIn, event.MouseButtonPress):
            self.target.setFocus()
        return super().eventFilter(obj, event)

class BackEndClass(QtWidgets.QWidget, Ui_MainWindow):

    def __init__(self):
        QtWidgets.QWidget.__init__(self)
        self.setupUi(MainWindow)
        self.tabWidget.tabBar().setVisible(False)
        self.user_btn.clicked.connect(self.user_mode)
        self.audit_btn.clicked.connect(self.audit_mode)
        self.back_btn_user.clicked.connect(self.back_menu)
        self.back_btn_audit.clicked.connect(self.back_menu)
        self.back_btn_user_2.clicked.connect(self.back_menu)
        self.pushButton_browse_audit.clicked.connect(self.browse_excel)
        self.pushButton_login.clicked.connect(self.login)
        self.lineEdit_password_audit_2.returnPressed.connect(self.login)
        self.FinishAudit.clicked.connect(self.finish_audit)
        self.ShowStatistics.clicked.connect(self.show_statistics)
        self.back_btn_stats.clicked.connect(self.backfromstats)
        self.new_login()
        global SCANNER_CONNECTED

        # Initialize Updater with the current version and repository details
        self.updater = Updater(
            current_version="v2.00",  # Replace with your tool's version
            repo_owner="ENGaliyasser",
            repo_name="JAKYA",
            progress_bar=self.progressBar
        )

        # Connect the update button to the updater's update function
        self.update.clicked.connect(self.updater.update_application)

        # Hide update-related UI elements initially
        self.progressBar.setVisible(False)
        self.update.setVisible(False)
        self.ask.setVisible(False)

        # Connect the "Check" button to the check_update function
        self.check.clicked.connect(lambda: self.updater.check_update(self.update, self.ask))
        
        # Access the live google sheet and save it in self.sheet
        try:
             # initalize the google sheet
            self.sheet = access_sheet.init_google_sheet()
        except:
            QMessageBox.about(self, "Message", "Can't access the google sheet. Hint: Check your internet connection.")

        # Find the USB serial device
        device_port = SCANNER.find_usb_serial_device()
        if device_port is None:
            print("No USB serial device found.")
            QMessageBox.about(self, "Message", "The Scanner is not Connected. Hint: Use Keyboard for Input")
            SCANNER_CONNECTED = False
        else:
            # update the scanner connected global variable
            SCANNER_CONNECTED = True
            # Set up the serial connection
            ser = SCANNER.serial.Serial(device_port, 9600)  # Adjust the baud rate as necessary
            # Create a QThread object
            self.thread = QThread()
            # Create a Scanner object
            self.scanner = SCANNER.Scanner(ser)
            # Move the scanner to the thread
            self.scanner.moveToThread(self.thread)
            # Connect signals and slots
            self.thread.started.connect(self.scanner.run)
            self.scanner.data_received.connect(self.scanner_data_update)
            app.aboutToQuit.connect(self.scanner.stop)
            app.aboutToQuit.connect(self.thread.quit)
            # Start the thread
            self.thread.start()

        # Functions for the hidden line Edit    
        if not SCANNER_CONNECTED:
            # The following functions use the hidden LineEdit to take input
            # Connect the signals with lambda to pass parameters
            self.lineEdit_audit.returnPressed.connect(lambda: self.insert_audit(1))
            self.scan_user.returnPressed.connect(lambda: self.Scan_user(1))

        # Class Attributes holding information for audit mode
        self.FailedTrials = 1         # Tracks the number of failed scans.
        self.SuccessTrials = 1         # Tracks the number of successful scans.
        self.ScannedAssets = 0         # Represents the total number of assets that have been scanned.
        self.RemainingAssets = 0       # Indicates the number of assets that are yet to be scanned.
        self.BSScannedAssets = 0       # Represents the number of Brightskies assets that have been scanned.
        self.BSremainingAssets = 0     # Indicates the number of Brightskies assets that are yet to be scanned.
        self.MahleScannedAssets = 0    # Represents the number of Mahle assets that have been scanned.
        self.MahleRemainingAssets = 0  # Indicates the number of Mahle assets that are yet to be scanned.

    # TODO: update the the login mode case
    def scanner_data_update(self, data):
        if CURRENT_MODE == "User Mode":
            self.Scan_user(data)
        elif CURRENT_MODE == "Audit Mode":
            self.insert_audit(data)
        elif CURRENT_MODE == "Login Mode":
            pass
        else:
            pass

    def Chart_init(self):
        # Create Matplotlib canvases for four pie charts
        print("hhhp")
        self.canvas_1 = pieChart.MplCanvas(self, width=8, height=6, dpi=100)
        self.canvas_2 = pieChart.MplCanvas(self, width=8, height=6, dpi=100)
        self.canvas_3 = pieChart.MplCanvas(self, width=8, height=6, dpi=100)
        self.canvas_4 = pieChart.MplCanvas(self, width=8, height=6, dpi=100)

        # Set up layouts to add the canvases directly into the widgets
        self.layout_1 = QVBoxLayout(self.widget)
        self.layout_2 = QVBoxLayout(self.widget_2)
        self.layout_3 = QVBoxLayout(self.widget_3)
        self.layout_4 = QVBoxLayout(self.widget_4)

        # Add the canvases to the respective layouts
        self.layout_1.addWidget(self.canvas_1)
        self.layout_2.addWidget(self.canvas_2)
        self.layout_3.addWidget(self.canvas_3)
        self.layout_4.addWidget(self.canvas_4)

        # Manually set the data (replace this with your data)
        labels_1 = ['Failed Trials ', 'Success Trials']
        data_1 = [self.FailedTrials, self.SuccessTrials]
        colors_1 = ['lightseagreen', 'red']  # Custom colors by name
        label_positions_1 = [(45, 1.5), (225, 1.5)]  # Example positions

        labels_2 = ['Scanned Assets', 'remaining Assets']
        data_2 = [self.ScannedAssets, self.RemainingAssets]
        colors_2 = ['lightseagreen', 'red']  # Custom colors by name
        label_positions_2 = [(45, 1.5), (225, 1.5)]  # Example positions

        labels_3 = ['Mahle Scanned Assets', 'Mahle remaining Assets']
        data_3 = [self.MahleScannedAssets, self.MahleRemainingAssets]
        colors_3 = ['papayawhip', 'grey']  # Custom colors by name
        label_positions_3 = [(45, 1.5), (225, 1.5)]  # Example positions

        labels_4 = ['BS Scanned Assets', 'BS remaining Assets']
        data_4 = [self.BSScannedAssets, self.BSremainingAssets]
        colors_4 = ['papayawhip', 'grey']  # Custom colors by name
        label_positions_4 = [(45, 1.5), (225, 1.5)]  # Example positions

        # Plot pie charts with custom colors
        self.canvas_1.plot_pie_chart(data_1, labels_1, colors_1, label_positions_1)
        self.canvas_2.plot_pie_chart(data_2, labels_2, colors_2, label_positions_2)
        self.canvas_3.plot_pie_chart(data_3, labels_3, colors_3, label_positions_3)
        self.canvas_4.plot_pie_chart(data_4, labels_4, colors_4, label_positions_4)


    def new_login(self):
        """
        Prepares the UI for a new login attempt by resetting fields and adjusting widget visibility.
        - Clears the password and Excel path fields.
        - Hides Excel-related widgets.
        - Shows password-related widgets.
        - Enables the password field and login button.
        """
        # Clear line edits
        self.lineEdit_password_audit_2.clear()
        self.lineEdit_excel_audit.clear()
        # Hide Excel related widgets
        self.label_excel.hide()
        self.lineEdit_excel_audit.hide()
        self.pushButton_browse_audit.hide()
        # Show password related widgets
        self.lineEdit_password_audit_2.show()
        self.pushButton_login.show()
        self.password_label_2.show()
        # enable login button and write password
        self.lineEdit_password_audit_2.setReadOnly(False)
        self.pushButton_login.setDisabled(False)
        self.Excel_Name = None
        self.right.hide() # Right Label for correct password


    def user_mode(self):
        """
        switches to user mode tab
        """
        # Switch to user mode tab
        self.tabWidget.setCurrentIndex(1)
        # Set focus to the hidden lineEdit and keep the focus on it
        self.scan_user.setFocus()
        self.scan_user.setCursor(QtCore.Qt.BlankCursor)
        self.scan_user.clear()
        # Remove other event filters (set focus of the audit mode tab)
        try:
            self.tabWidget.removeEventFilter(self.focus_redirector_audit)
        except:
            pass
        # Install the event filter on the tab widget in user mode (keep focus on the lineEdit in user mode)
        self.focus_redirector_user = FocusRedirector(self.scan_user)
        self.tabWidget.installEventFilter(self.focus_redirector_user)
        # Set current mode to user mode
        global CURRENT_MODE
        CURRENT_MODE = "User Mode"

    def audit_mode(self):
        """
        Switches to audit mode login tab
        """
        self.tabWidget.setCurrentIndex(2)
        # Change current mode to Login Mode
        global CURRENT_MODE
        CURRENT_MODE = "Login Mode"

    def back_menu(self):
        """
        Returns back to the main menu
        """
        self.tabWidget.setCurrentIndex(0)
        self.new_login()
        global CURRENT_MODE
        CURRENT_MODE = None


    def login(self):
        """
        Handles user login by checking the entered password and updating the UI accordingly.

        - If the password is correct:
            - Disables further password input.
            - Shows Excel-related widgets.
            - Displays a success message.
        - If the password is incorrect:
            - Displays an error message.
        """
        # Get the entered password from the line edit
        entered_password = self.lineEdit_password_audit_2.text()
        # Password Check
        if check_password(entered_password):
            # Disable the line edit and the login button
            self.lineEdit_password_audit_2.setReadOnly(True)
            self.pushButton_login.setDisabled(True)
            # Show the right sign
            self.right.show()
            # Show Excel related widgets
            self.label_excel.show()
            self.lineEdit_excel_audit.show()
            self.pushButton_browse_audit.show()
        else:
            QMessageBox.about(self, "Message", "Invalid password! Please Enter the Correct Password.")

    def browse_excel(self):
        """
        Opens a file dialog for selecting an Excel file and updates the UI based on the selection.
        - Prompts the user to select an Excel (.xlsx) file.
        - If a valid file is selected:
            - Stores the file path.
            - Displays the path in the UI.
            - If the file format is correct, shows a success message and starts the audit.
            - If the format is incorrect, shows an error message.
        - If no file is selected, shows an error message.
        """
        Excel_File = QFileDialog.getOpenFileName(self, 'Open File', 'Select File', '(*.xlsx)')
        Excel_File = Excel_File[0]
        if Excel_File.endswith(".xlsx"):
            self.Excel_Name = Excel_File
            self.lineEdit_excel_audit.setText(Excel_File)
            if self.check_excel_format(Excel_File):
                # Text box for lets start the audit
                # self.msg_box = QMessageBox(self)
                # self.msg_box.setWindowTitle("Success")
                # self.msg_box.setText("Let's start the audit")
                # self.msg_box.show()
                # Create a QTimer to close the message box after 2 seconds
                #QTimer.singleShot(2000, self.enter_the_audit)
                self.enter_the_audit()
            else:
                QMessageBox.about(self, "Message", "Invalid file format, please upload a valid excel file.")
        else:
            QMessageBox.about(self, "Message", "No File selected!")

    def enter_the_audit(self):
        """
        Closes the message box and switches to the audit tab.
        """
        # Close the QmessageBox
        # self.msg_box.close()
        # Go to the audit tab
        self.tabWidget.setCurrentIndex(3)
        self.Display_audit()
        # Show the audit name and date in the labels
        self.textBrowser_5.setText(str(date.today()))
        self.textBrowser_4.setText(os.path.basename(str(self.Excel_Name)))

        # Set focus on the hidden LineEdit
        self.lineEdit_audit.setFocus()
        self.lineEdit_audit.setCursor(QtCore.Qt.BlankCursor)
        # Remove other event filters (user LineEdit focus)
        try:
            self.tabWidget.removeEventFilter(self.focus_redirector_user)
        except:
            pass
        # Install the event filter on the tab widget in audit mode
        self.focus_redirector_audit = FocusRedirector(self.lineEdit_audit)
        self.tabWidget.installEventFilter(self.focus_redirector_audit)
        # Switch Current mode to Audit Mode
        global CURRENT_MODE
        CURRENT_MODE = "Audit Mode"
        

    def insert_audit(self, data):
        """
           - Retrieves the asset tag entered by the user from the input field.
           - Attempts to open 'Audit_Output.xlsx' for updates; if not available, opens the original Excel file.
           - Searches for the asset tag in column 2 of the sheet.
           - If the asset tag is found, marks "True" in column 6 of the corresponding row.
           - Saves the updated data to 'Audit_Output.xlsx' and updates the display.
           - If the asset tag is not found, shows an error message.
           - Catches and handles any errors that occur during the process.
           """
        # Clear the Start scanning hint label
        self.start_scan_label.clear()

        # Get the asset tag
        if SCANNER_CONNECTED:
            AssetTag = str(data) # directly from the signal passed when a scan is made.
        else:
            AssetTag = str(self.lineEdit_audit.text()) # from the hidden line edit.

        print("Asset Tag: " + AssetTag)
        try:
            try:
                wb = openpyxl.load_workbook("Audit_Output.xlsx")
            except Exception as e:
                wb = openpyxl.load_workbook(self.Excel_Name)
            sheet = wb.active
            row_to_update = None

            for i in range(1, sheet.max_row + 1):
                if str(sheet.cell(row=i, column=self.asset_tag_col).value) == AssetTag:
                    # Mark this row as "True" in column 6
                    if sheet.cell(row=i, column=self.status_col).value == "True":
                        self.start_scan_label.setText("Item Already Scanned!")
                    else :
                        self.RemainingAssets -= 1
                        self.ScannedAssets += 1
                        if sheet.cell(row=i, column=self.owner_col).value.lower() == "brightskies":
                            self.BSremainingAssets -= 1
                            self.BSScannedAssets += 1
                        elif sheet.cell(row=i, column=self.owner_col).value.lower() == "brightskies":
                            self.MahleRemainingAssets -= 1
                            self.MahleScannedAssets += 1
                    sheet.cell(row=i, column=self.status_col).value = "True"
                    row_to_update = i
                    break

            if row_to_update:
                wb.save("Audit_Output.xlsx")
                self.Display_audit()
                self.SuccessTrials += 1
            else:
                if not self.lineEdit_audit.text() == "":
                    QMessageBox.about(self, "Message", "Asset Tag not found in the Excel sheet")
                    self.FailedTrials += 1

        except Exception as e:
            QMessageBox.about(self, "Message", "Error: " + str(e))
        self.lineEdit_audit.clear()
        self.lineEdit_audit.setFocus()

    def Display_audit(self):
        """
            - Attempts to open 'Audit_Output.xlsx'; if not found, opens the original Excel file.
            - Displays the total number of rows in the sheet in 'textBrowser_6'.
            - Initializes counters for marked ("True") and unmarked rows.
            - Iterates through the rows to count how many rows are marked as "True" in column 6.
            - Updates 'textBrowser_14' and 'textBrowser_16' with the count of marked rows.
            - Updates 'textBrowser_15' with the count of unmarked rows.
            """
        try:
            wb = openpyxl.load_workbook("Audit_Output.xlsx")
        except Exception as e:
            wb = openpyxl.load_workbook(self.Excel_Name)
        sheet = wb.active

        self.textBrowser_6.setText(str(sheet.max_row))
        self.textBrowser_13.setText("0")
        self.textBrowser_14.setText("0")
        k = 0
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=self.status_col).value == "True":
                k += 1
        self.textBrowser_14.setText(str(k))
        self.textBrowser_15.setText(str(sheet.max_row - k))
        self.textBrowser_16.setText(str(k))

    def Scan_user(self, data):
        """
            - Retrieves the asset tag entered by the user from the input field.
            - Opens the live google sheet excel to check for the asset tag.
            - If a matching asset tag is found, it updates several text browsers with values from that row.

            """
        # Reset Labels on each new scan
        self.asset_name.clear()
        self.asset_tag.clear()
        self.serial.clear()
        self.checked_out_to.clear()
        self.owner.clear()

        # Asset Tag
        if SCANNER_CONNECTED:
            AssetTag = data # Directly from the scanner
        else:
            AssetTag = str(self.scan_user.text()) # From the hidden text box
        

        print("Asset Tag: " + AssetTag)
        try:
            all_values = self.sheet.get_all_values()
            # get the headers from the first row
            headers = all_values[0]
            print("Headers:", headers)
            # Determine column indexes
            self.asset_tag_col = headers.index('Asset Tag') + 1
            self.serial_col = headers.index('Serial') + 1
            self.check_out_to_col = headers.index('checked out to') + 1
            self.owner_col = headers.index('owner') + 1
            self.asset_name_col = headers.index('Asset Name') + 1

            # Search for the asset tag
            row_to_update = None
            for row in all_values[1:]:  # Skip header row
                if row[self.asset_tag_col - 1] == AssetTag:
                    self.asset_name.setText(row[self.asset_name_col - 1])
                    self.asset_tag.setText(row[self.asset_tag_col - 1])
                    self.serial.setText(row[self.serial_col - 1])
                    self.checked_out_to.setText(row[self.check_out_to_col - 1])
                    self.owner.setText(row[self.owner_col - 1])
                    row_to_update = row
                    break
            if row_to_update == None:
                QMessageBox.about(self, "Message", "Asset Tag not found in the Excel sheet")

        except Exception as e:
            QMessageBox.about(self, "Message", "Can't Access the Google Sheet, Hint: Check your internet connection. Error: " + str(e))
        self.scan_user.clear()

    def check_excel_format(self, excel_file):
        """Checks if the excel file has the correct format.

        Args:
            excel_file: Path to the Excel file.

        Returns:
            True if the sheet format is correct, False otherwise.
        """
        print("jj")
        if(excel_file != os.getcwd().replace("\\", "/") + "/Audit_Output.xlsx"):
            try:
                print("here464")
                workbook = openpyxl.load_workbook("Audit_Output.xlsx")
                QMessageBox.about(self, "Error", "Please move the Audit_Output.xlsx file to a different folder, or choose it as the Excel file.")
                return False
            except Exception as e:
                print(e)
                workbook = openpyxl.load_workbook(excel_file)
                print("kk")
        else:
            workbook = openpyxl.load_workbook("Audit_Output.xlsx")

        sheet = workbook.active
        # defining the excpected columns
        # TODO:
        # Add Model, Category, Status, Checked Out To, Location, Purchase Cost, HS-Code, Owner
        # Important (ERROR): Asset Name, Asset Tag, Serial, Checked Out To, Owner (USER MODE)
        # Warning: The rest
        expected_columns = set(["asset name", "asset tag", "serial", "checked out to", "model", "category", "status",
                            "location", "purchase cost", "hs-code", "owner"])
        important_columns = set(["asset name", "asset tag", "serial", "checked out to", "owner"])

        # Get the actual number of columns in the first row
        print(sheet.max_row)
        print(sheet.max_column)

        actual_columns = set()

        for i in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=i).value
            if cell_value is not None:
                actual_columns.add(cell_value.lower())

        print("cc")
        print(actual_columns)
        print(important_columns)

        # Check if the column titles match the expected titles
        if not important_columns.issubset(actual_columns):
            return False
        # All correct Return True
        flag_status = False
        print("ff")
        for i in range(1, sheet.max_column + 1):
            if (sheet.cell(row = 1, column = i).value) is None:
                continue
            if (sheet.cell(row = 1, column = i).value).lower() == "asset tag":
                self.asset_tag_col = i
            if (sheet.cell(row = 1, column = i).value).lower() == "status":
                self.status_col = i
                flag_status = True
            if (sheet.cell(row = 1, column = i).value).lower() == "serial":
                self.serial_col = i
                print(self.serial_col)
            if (sheet.cell(row = 1, column = i).value).lower() == "check out to":
                self.check_out_to_col = i
            if (sheet.cell(row = 1, column = i).value).lower() == "owner":
                self.owner_col = i
            if (sheet.cell(row = 1, column = i).value).lower() == "asset name":
                self.asset_name_col = i


        print("4")
        if not flag_status:
            self.status_col = sheet.max_row + 1
            sheet.cell(row = 1, column = self.status_col).value = "Status"
        print("5")
        workbook.save("Audit_Output.xlsx")
        try:
            self.InitStats()
        except Exception as e:
            print(e)
        return True

    def finish_audit(self):
         # TODO: use this function to generate any required files Execl, pie charts, reports
         # Format the data
        data = f"Date: {str(date.today())} \n" \
            f"Audit: {os.path.basename(str(self.Excel_Name))} \n" \
            f"Scanned Assets: {self.ScannedAssets} \n" \
            f"Remaining Assets: {self.RemainingAssets} \n" \
            f"Brightskies Scanned Assets: {self.BSScannedAssets} \n" \
            f"Brightskies Remaining Assets: {self.BSremainingAssets} \n" \
            f"Mahle Scanned Assets: {self.MahleScannedAssets} \n" \
            f"Mahle Remaining Assets: {self.MahleRemainingAssets} \n"
        
        # Use the createFiles module to create the results folder
        if createfiles.create_folder("Results"):
            createfiles.create_textfile("Results","results.txt", data)
        else:
            answer = QMessageBox.question(self,'', "Results Already Exist, overwrite?", QMessageBox.Yes | QMessageBox.No)
            if answer == QMessageBox.Yes:
                createfiles.delete_folder("Results")
                createfiles.create_textfile("Results","results.txt", data)
            elif answer == QMessageBox.No:
                pass

    def show_statistics(self):
        try:
            self.Chart_init()
        except Exception as e:
            print(e)
        self.tabWidget.setCurrentIndex(4)

    def backfromstats(self):
        self.tabWidget.setCurrentIndex(3)
        
    def InitStats(self):
        print("hah")
        workbook = openpyxl.load_workbook("Audit_Output.xlsx")
        print("hoh")
        sheet = workbook.active
        print("hih")
        try:
            self.RemainingAssets = sheet.max_row
        except Exception as e:
            print(e)

        print("haha")

        for i in range(1, sheet.max_row + 1):
            if (sheet.cell(row = i, column = self.owner_col).value) is None:
                continue
            if (sheet.cell(row = i, column = self.owner_col).value).lower() == "brightskies":
                self.BSremainingAssets += 1
            if (sheet.cell(row = i, column = self.owner_col).value).lower() == "mahle":
                self.MahleRemainingAssets += 1

        print("hehe")




if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    Updater.delete_old_versions(current_version="v2.00")  # Replace with your tool's version
    MainWindow = QtWidgets.QMainWindow()
    ui = BackEndClass()
    MainWindow.show()
    sys.exit(app.exec_())
